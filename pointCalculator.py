import json
import openpyxl
import math

# Load the JSON file
with open('coefficients.json') as f:
    data = json.load(f)
    print("JSON Loaded")

def calculateResult(sex, event, points):
    coeffs = data[sex][event]
    resultShift = coeffs["resultShift"]
    conversionFactor = coeffs["conversionFactor"]
    pointShift = coeffs["pointShift"]
    
    # Calculating the time/distance from the points
    if event in ["High Jump", "Pole Vault", "Long Jump", "Triple Jump", "Javelin", "Shot Put", "Discus", "Hammer"]:
        result = -resultShift + math.sqrt((points - pointShift) / conversionFactor)
    else:
        result = -resultShift - math.sqrt((points - pointShift) / conversionFactor)
    
    return result

def calculatePoints(sex, event, result):
    coeffs = data[sex][event]
    resultShift = coeffs["resultShift"]
    conversionFactor = coeffs["conversionFactor"]
    pointShift = coeffs["pointShift"]

    # Calculating the points from the time/distance
    points = conversionFactor * ((result + resultShift) * (result + resultShift)) + pointShift
    return points

def insert_into_excel():
    # Load the Excel file
    workbook = openpyxl.load_workbook('Cleaned-Data.xlsx')
    print("Workbook Loaded")
    sheet = workbook.active
    print("Sheet Loaded")

    rownum = 2

    # Iterate through rows
    for row in sheet.iter_rows(min_row=rownum, values_only=True):  # Assuming data starts from row 2
        sex = row[3]  # Pull the sex from column D
        event = row[6]  # Pull the event from column G
        result = row[1] # Pull the result from column B
        #print("Reading row " + str(row))

        # Access the data using keys from the Excel file
        if sex in data and event in data[sex]:
            score = calculatePoints(sex, event, result)

            # Write the score to column K
            sheet.cell(row=rownum, column=11).value = score  # Column K corresponds to index 11

        else:
            print(f"Data not found for {sex} {event}")
            print(row)

        rownum = rownum + 1     # Increment the row number for the next iteration

    # Save the modified Excel file
    workbook.save('Cleaned-Data.xlsx')