import openpyxl
import pandas as pd

df = pd.read_csv('Cleaned-Data.csv')

# Load the Excel file
workbook = openpyxl.load_workbook('top100avg.xlsx')
print("Workbook Loaded")
sheet = workbook.active
print("Sheet Loaded")

events = ['100m', '200m', '400m', '800m', '1500m', '1600m', '3200m', '5000m', '10000m', 'Triple Jump', 'High Jump', 'Long Jump', 'Shot Put', 'Discus', 'Javelin']  # Add the events to include
divisions = ['NCAA D-I', 'NCAA D-II', 'NCAA D-III', 'NAIA', 'Kansas 1A', 'Kansas 3A', 'Kansas 6A', 'World']  # Add the divisions to include
years = [2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019, 2021, 2022, 2023]  # Add the years to include
genders = ['Men', 'Women']

rownum = 1
yearcol = 4
for event in events:
    event_df = df[df['Event'] == event]
    for division in divisions:
        division_df = event_df[event_df['Division'] == division]
        for gender in genders:
            gender_df = division_df[division_df['Sex'] == gender]
            for year in years:
                year_df = gender_df[gender_df['Year'] == year]
                average_points = year_df['Points'].mean()
                sheet.cell(row=rownum, column=1).value = division
                sheet.cell(row=rownum, column=2).value = gender
                sheet.cell(row=rownum, column=3).value = event
                sheet.cell(row=rownum, column=yearcol).value = average_points
                yearcol = yearcol + 1
                if year == 2023:
                    yearcol = 4
            rownum = rownum + 1

workbook.save('top100avg.xlsx')